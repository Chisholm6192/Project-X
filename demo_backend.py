from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import math
storage = {}
headings = []

wb = Workbook()
ws = wb.active
class sprout:
    def __init__(self,name):
        self.name = name
        if os.path.isfile(f'{name}.xlsx') == True:
            self.profile = load_workbook(f'{name}.xlsx')
        else:
            wb.title = name
            ws.append(['Credit card info','Amount', 'Donation', 'Total Contribuations'])
            wb.save(f'{name}.xlsx')
    def check_cells(self,Credit=False,money=False,donation=False):
        self.creditinfo = []
        self.money = []
        self.donation = []
        for row in range(1000):
            if Credit == True:
                self.creditinfo.append(ws[get_column_letter(1)+str(row+1)].value)
                if ws[get_column_letter(1)+str(row+1)].value == '':
                    break
            if money == True:
                self.creditinfo.append(ws[get_column_letter(2)+str(row+1)].value)
                if ws[get_column_letter(2)+str(row+1)].value == '':
                    break
            if donation == True:
                self.creditinfo.append(ws[get_column_letter(3)+str(row+1)].value)
                if ws[get_column_letter(2)+str(row+1)].value == '':
                    break
    def donationCalc(self, money, total=False):
        if total == False:
            dono = math.ceil(money) - money
            self.donation.append(dono)
        else:
            dono = sum(self.donation)
        return dono
    def addEntry(self, credit,money):
        self.check_cells(Credit=True)
        self.check_cells(money=True)
        self.check_cells(donation=True)
        ws[get_column_letter(1)+'2'] = credit
        ws.column_dimensions[get_column_letter(1)].width = len(credit)
        ws[get_column_letter(2)+'2'] = money
        ws[get_column_letter(3)+'2'] = self.donationCalc(money)
        ws[get_column_letter(4)+'2'] = self.donationCalc(money,total=True)
        wb.save(f'{self.name}.xlsx')


profile = sprout('Aaron')
print(profile.addEntry('0123 0124 0124 0543', 14.23))

